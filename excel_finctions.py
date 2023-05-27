import openpyxl
import os

workbook = openpyxl.load_workbook('example.xlsx')
#get all sheets
sheets = workbook.sheetnames
print(sheets)
#load sheet
sheet = workbook['Sheet1']
#get known cell value
print(sheet['B1'].value)

#rows and columns start at 1, not 0
for i in range(1,7):
    print(i, sheet.cell(row=i, column=2).value)

#create new workbook
new_workbook_name = 'created_by_code.xlsx'
if os.path.exists(new_workbook_name):
    os.remove(new_workbook_name)

new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.create_sheet('Sheet')
new_sheet['A1'] = 42
sheet2 = new_workbook.create_sheet()
sheet2.title = 'Bungalo'
new_workbook.create_sheet(index=0, title='First Sheet')
new_workbook.save(new_workbook_name)